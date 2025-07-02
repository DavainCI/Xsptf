import tkinter as tk
from tkinter import messagebox
import spotipy
from spotipy.oauth2 import SpotifyClientCredentials
import os
import requests
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image as PILImage

# Spotify color palette 
# Para tkinter (con #)
SPOTIFY_GREEN_TK = "#1DB954"
DARK_GRAY_TK = "#191414"
LIGHT_GRAY_TK = "#EFEFEF"

# Para openpyxl (con FF al inicio para aRGB)
SPOTIFY_GREEN_XL = "FF1DB954"
DARK_GRAY_XL = "FF191414"
LIGHT_GRAY_XL = "FFEFEFEF"

# Output folder
OUTPUT_FOLDER = "PlaylistsExcel"
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def clean_filename(name):
    """Clean playlist name to make it safe for filenames"""
    invalid_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
    for char in invalid_chars:
        name = name.replace(char, "_")
    return name

def show_custom_popup(message, excel_path):
    """Show custom popup with options to open file/folder"""
    root = tk.Toplevel() 
    root.title("✅ Playlist Exportada")
    root.geometry("500x250")  # Aumenté un poco el tamaño
    window_width = 500
    window_height = 250
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_x = int(screen_width/2 - window_width/2)
    position_y = int(screen_height/2 - window_height/2)
    root.geometry(f"{window_width}x{window_height}+{position_x}+{position_y}")
    root.configure(bg=DARK_GRAY_TK)
    
    # Frame principal con bordes redondeados
    main_frame = tk.Frame(
        root, 
        bg=DARK_GRAY_TK,
        highlightbackground=SPOTIFY_GREEN_TK,
        highlightthickness=2,
        bd=0,
        relief='solid'
    )
    main_frame.pack(expand=True, fill='both', padx=20, pady=20)
    
    # Main message - Formato mejorado
    msg_frame = tk.Frame(main_frame, bg=DARK_GRAY_TK)
    msg_frame.pack(pady=(20, 10), padx=20, fill='both', expand=True)
    
    lines = message.split('\n')
    for line in lines:
        tk.Label(
            msg_frame, 
            text=line,
            bg=DARK_GRAY_TK,
            fg=LIGHT_GRAY_TK if not line.startswith("📌") else SPOTIFY_GREEN_TK,
            font=("Arial", 11, "bold" if line.startswith("Playlist") else "normal"),
            anchor='w'
        ).pack(fill='x', pady=2)
    
    # Button frame
    btn_frame = tk.Frame(main_frame, bg=DARK_GRAY_TK)
    btn_frame.pack(pady=(10, 20))
    
    # Función para crear botones redondeados
    def create_rounded_button(parent, text, command, bg_color, fg_color):
        btn = tk.Button(
            parent,
            text=text,
            command=command,
            bg=bg_color,
            fg=fg_color,
            font=("Arial", 10, "bold"),
            bd=0,
            highlightthickness=0,
            relief='flat',
            padx=15,
            pady=5
        )
        # Simular bordes redondeados (tkinter no los soporta nativamente)
        btn.config(activebackground=bg_color, activeforeground=fg_color)
        btn.bind("<Enter>", lambda e: btn.config(bg=bg_color))
        btn.bind("<Leave>", lambda e: btn.config(bg=bg_color))
        return btn
    
    # Buttons with rounded style
    create_rounded_button(
        btn_frame, "Abrir Archivo", 
        lambda: os.startfile(excel_path),
        SPOTIFY_GREEN_TK, "white"
    ).pack(side=tk.LEFT, padx=10)
    
    create_rounded_button(
        btn_frame, "Abrir Carpeta", 
        lambda: os.startfile(os.path.dirname(excel_path)),
        "#333333", "white"
    ).pack(side=tk.LEFT, padx=10)
    
    create_rounded_button(
        btn_frame, "Cerrar", 
        root.destroy,
        "#555555", "white"
    ).pack(side=tk.LEFT, padx=10)

    root.grab_set()
    root.wait_window() 

def main():
    try:
        # Spotify API credentials
        client_id = "71114e96572a4b759750f90f89653e12"
        client_secret = "44374ebc9731491e87bee7fad0156a2c"

        # Authenticate with Spotify
        auth_manager = SpotifyClientCredentials(client_id=client_id, client_secret=client_secret)
        sp = spotipy.Spotify(auth_manager=auth_manager)

        # Get playlist URL from user
        playlist_url = input("🎵 Pega la URL de la playlist de Spotify: ").strip()
        playlist_id = playlist_url.split("/")[-1].split("?")[0]

        # Get playlist data
        playlist = sp.playlist(playlist_id)
        playlist_name = playlist['name']
        playlist_image_url = playlist['images'][0]['url']
        print(f"📌 Playlist: {playlist_name}")
        print(f"🌉 URL de la imagen: {playlist_image_url}")
        
        # Create safe filename
        safe_playlist_name = clean_filename(playlist_name)
        excel_file = os.path.join(OUTPUT_FOLDER, f"{safe_playlist_name}.xlsx")
        img_path = os.path.join(OUTPUT_FOLDER, f"{safe_playlist_name}_image.png")

        # Download playlist image
        response = requests.get(playlist_image_url)
        if response.status_code != 200:
            raise Exception(f"Error al descargar la imagen (Código {response.status_code})")
        
        with open(img_path, 'wb') as f:
            f.write(response.content)

        # Process tracks
        tracks_data = []
        for item in playlist['tracks']['items']:
            track = item['track']
            name = track['name']
            url = track['external_urls']['spotify']
            artist = ", ".join([a['name'] for a in track['artists']])
            duration_ms = track['duration_ms']
            duration = f"{duration_ms//60000}:{(duration_ms%60000)//1000:02d}"
            tracks_data.append((name, artist, duration, url))

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Playlist"

        # Add headers with Spotify style
        headers = ["Canción", "Artista(s)", "Duración"]
        ws.append(headers)
        
        header_style = {
            'font': Font(bold=True, color="FFFFFF", name="Arial"),
            'fill': PatternFill("solid", fgColor=SPOTIFY_GREEN_XL),  # Usamos la versión para Excel
            'alignment': Alignment(horizontal="center", vertical="center")
        }
        
        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            for attr, value in header_style.items():
                setattr(cell, attr, value)

        # Add tracks data
        for row_idx, (name, artist, duration, url) in enumerate(tracks_data, start=2):
            # Song name with hyperlink
            ws.cell(row=row_idx, column=1, value=name).hyperlink = url
            ws.cell(row=row_idx, column=1).font = Font(name="Arial", color="0563C1", underline="single")
            
            # Artist
            ws.cell(row=row_idx, column=2, value=artist).font = Font(name="Arial", size=11)
            
            # Duration (centered)
            ws.cell(row=row_idx, column=3, value=duration)
            ws.cell(row=row_idx, column=3).font = Font(name="Arial", size=11)
            ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal="center")

        # Set column widths
        ws.column_dimensions["A"].width = 38.5
        ws.column_dimensions["B"].width = 38.5
        ws.column_dimensions["C"].width = 12

        # Add playlist image (resized)
        img = PILImage.open(img_path).resize((150, 150))
        img.save(img_path)
        ws.add_image(XLImage(img_path), "E2")

        # Add playlist title
        title_cell = ws.cell(row=10, column=5, value=f"Playlist: {playlist_name}")
        title_cell.font = Font(name="Arial", size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        # Save the workbook
        wb.save(excel_file)
        print(f"\n💾 Archivo guardado en: {os.path.abspath(excel_file)}")
        print(f"🎵 Total de canciones: {len(tracks_data)}")

        # Show success popup
        show_custom_popup(
        f"Playlist exportada con éxito!\n\n"
        f"📌 Nombre: {playlist_name}\n"
        f"🎵 Canciones: {len(tracks_data)}\n"
        f"📁 Ubicación: {os.path.abspath(excel_file)}",
        excel_file
)

    except Exception as e:
        # Show error message
        error_root = tk.Tk()
        error_root.withdraw()
        messagebox.showerror(
            "❌ Error",
            f"No se pudo exportar la playlist:\n\n{str(e)}"
        )
        print(f"\n❌ Error: {str(e)}")

if __name__ == "__main__":
    main()